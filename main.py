"""
DevSuite — FastAPI Backend  (v2.2.0)
---------------------------------
Serves the static frontend and provides REST/WebSocket APIs for all tools.

Persistence
-----------
All data is stored through the DevDB engine in a single KeePass-style
binary file at ~/.devsuite/devdb.dsb.  The server never decrypts the
client-side AES-encrypted blobs inside the vault/ssh_profiles stores."""

import asyncio
import ipaddress
import json
import logging
import os
import re
import secrets
import socket
import stat
import string
import struct
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Annotated

import asyncssh
from fastapi import (
    FastAPI,
    File,
    Form,
    HTTPException,
    Request,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from devdb import DevDB

# PTY support is Linux/macOS only — import conditionally so the module loads on Windows.
pty = None  # pylint: disable=invalid-name
fcntl = None  # pylint: disable=invalid-name
termios = None  # pylint: disable=invalid-name
_pty_available = False
if sys.platform != 'win32':
    try:
        import pty  # type: ignore[assignment]
        import fcntl  # type: ignore[assignment]
        import termios  # type: ignore[assignment]
        _pty_available = True
    except ImportError:
        pass

_PTY_AVAILABLE = _pty_available  # compatibility alias for tests referencing main._PTY_AVAILABLE

logger = logging.getLogger("devsuite")

# ─── DevDB — Unified Encrypted Database ────────────────────────────────────────
_DEVSUITE_DIR  = Path.home() / ".devsuite"
_DB_PATH       = _DEVSUITE_DIR / "devdb.dsb"
_LEGACY_URL_DB = Path(__file__).parent / "url_db.json"   # in-repo legacy file

_db = DevDB(_DB_PATH, password=os.environ.get("DEVDB_PASSWORD") or None)


@asynccontextmanager
async def _lifespan(_application: FastAPI):
    """Modern FastAPI lifespan: open DevDB, migrate legacy files, seed url_db cache."""
    _db.open()
    migrated = DevDB.migrate_legacy(_db, _DEVSUITE_DIR, _LEGACY_URL_DB)
    if migrated:
        _db.save()
        logger.info("DevDB: migration complete, database saved to %s", _DB_PATH)
    else:
        logger.info("DevDB: opened %s (%d bytes)", _DB_PATH, _db.file_size())
    url_db.update(_db.get_store("url_db"))
    yield  # app is running
    # (cleanup goes here if needed in future)


APP_VERSION = "0.1.3"

app = FastAPI(
    title="DevSuite",
    description="A private, locally-hosted developer suite with encrypted unified storage.",
    version=APP_VERSION,
    lifespan=_lifespan,
)

# Allowlist of hostnames that the /api/proxy endpoint is permitted to contact.
# Adjust this set to match the APIs you intend to test via the proxy.
ALLOWED_PROXY_HOSTS = {
    # Example public APIs:
    "api.github.com",
    "jsonplaceholder.typicode.com",
    "httpbin.org",
}

# ─── Shared string constants ──────────────────────────────────────────────────
_ALLOWED_ORIGINS      = ["http://localhost:8000", "http://127.0.0.1:8000"]  # NOSONAR — localhost-only CORS; HTTPS not applicable for loopback dev server
_ERR_ORIGIN_REQUIRED  = "Origin header required"
_ERR_ORIGIN_NOT_ALLOWED = "Origin not allowed"
_ERR_SFTP_FAILED      = "SFTP operation failed"
_OPENPYXL_MISSING     = "openpyxl is not installed. Run: pip install openpyxl"
_MIME_OCTET_STREAM    = "application/octet-stream"
_RE_NON_DIGIT         = r'\D'
_WSL_EXE              = "wsl.exe"

static_dir = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(static_dir, exist_ok=True)

# Serve static assets (JS, CSS, images) from the /static route
app.mount("/static", StaticFiles(directory=static_dir), name="static")

# ─── URL Shortener helpers (backed by DevDB 'url_db' store) ────────────────────


def _load_url_db() -> dict:
    """Load the url_db store from DevDB (falls back to empty dict)."""
    return _db.get_store("url_db")


def _save_url_db(data: dict) -> None:
    """Persist the url_db store to DevDB."""
    _db.set_store("url_db", data)
    _db.save()


# In-memory cache for the URL shortener (populated at startup via _startup_devdb)
# This proxy ensures the rest of the shortener code has the dict in scope.
url_db: dict = {}

# Maximum size for file uploads to /api/convert (20 MB)
MAX_UPLOAD_SIZE = 20 * 1024 * 1024
_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ShortenRequest(BaseModel):
    """Request body for the /api/shorten endpoint."""

    url: str


@app.middleware("http")
async def add_security_headers(request, call_next):
    """
    Attach a standard set of HTTP security headers to every outgoing response.

    Designed for use as FastAPI HTTP middleware: invokes the downstream handler and augments
    the returned response with headers that mitigate clickjacking, MIME-type sniffing,
    some XSS vectors, enforce HSTS, and provide a restrictive Content Security Policy.

    Parameters:
        request: The incoming ASGI/Starlette request object.
        call_next: A callable that accepts the request and returns a response from the
            downstream route/handler.

    Returns:
        The downstream response with the security headers added.
    """
    response = await call_next(request)
    # Prevent clickjacking
    response.headers["X-Frame-Options"] = "DENY"
    # Prevent MIME-type sniffing
    response.headers["X-Content-Type-Options"] = "nosniff"
    # Basic XSS protection for older browsers
    response.headers["X-XSS-Protection"] = "1; mode=block"
    # Content Security Policy (allows local assets and CDN resources used in the app)
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' blob: "
        "https://cdnjs.cloudflare.com https://cdn.jsdelivr.net; "
        "worker-src 'self' blob:; "
        "style-src 'self' 'unsafe-inline' "
        "https://fonts.googleapis.com https://cdnjs.cloudflare.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self';"
    )
    response.headers["Content-Security-Policy"] = csp
    return response


_STATIC_ASSET_RE = re.compile(r'(/static/[^"\'?]+\.(?:css|js))(?:\?v=[^"\']*)?')
_FAVICON_TAG = (
    '<link rel="icon" href="/static/favicon.svg" type="image/svg+xml">\n'
    '    <link rel="icon" href="/static/favicon.svg" sizes="any">'
)


def _serve_html(filename: str) -> str:
    """Read an HTML file, inject the favicon and cache-busting version into all local static asset URLs."""
    html_path = os.path.join(static_dir, filename)
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"{filename} not found.") from None  # NOSONAR — helper raises documented at call-site routes
    # Inject favicon into <head> if not already present
    if 'favicon' not in html:
        html = html.replace('<head>', f'<head>\n    {_FAVICON_TAG}', 1)
    return _STATIC_ASSET_RE.sub(rf'\1?v={APP_VERSION}', html)


@app.get("/", response_class=HTMLResponse, summary="Serve DevSuite homepage")
def read_home():
    """Serve the DevSuite landing page."""
    return _serve_html("home.html")


@app.get("/tools", response_class=HTMLResponse, summary="Serve DevSuite tools page")
def read_tools():
    """Serve the DevSuite tools dashboard (all tools grid)."""
    return _serve_html("tools.html")


@app.get("/diff", response_class=HTMLResponse, summary="Serve diff tool")
def read_diff():
    """Serve the Text/Folder Diff tool."""
    return _serve_html("index.html")


@app.get("/json", response_class=HTMLResponse, summary="Serve JSON linter tool")
def read_json_tool():
    """Serve the JSON Linter & Formatter tool."""
    return _serve_html("json.html")


@app.get("/yaml", response_class=HTMLResponse, summary="Serve YAML linter tool")
def read_yaml_tool():
    """Serve the YAML Linter & Validator tool."""
    return _serve_html("yaml.html")


@app.get("/regex", response_class=HTMLResponse, summary="Serve Regex Tester tool")
def read_regex_tool():
    """Serve the Regex Tester tool."""
    return _serve_html("regex.html")


@app.get("/base64", response_class=HTMLResponse, summary="Serve Base64 Encoder/Decoder tool")
def read_base64_tool():
    """Serve the Base64 Encoder/Decoder tool."""
    return _serve_html("base64.html")


@app.get("/crypto", response_class=HTMLResponse, summary="Serve Crypto Suite tool")
def read_crypto_tool():
    """Serve the Crypto Suite tool (Hash, AES, RSA, HMAC)."""
    return _serve_html("crypto.html")


@app.get("/url-shortener", response_class=HTMLResponse, summary="Serve URL Shortener tool")
def read_url_shortener_tool():
    """Serve the Link & QR Studio tool."""
    return _serve_html("url-shortener.html")


@app.get("/api-tester", response_class=HTMLResponse, summary="Serve Local API Tester tool")
def read_api_tester_tool():
    """Serve the API Tester tool."""
    return _serve_html("api-tester.html")


@app.get("/ssh", response_class=HTMLResponse, summary="Serve SSH & SFTP Manager tool")
def read_ssh_tool():
    """Serve the SSH & SFTP Manager tool."""
    return _serve_html("ssh-manager.html")


@app.get("/sftp", response_class=HTMLResponse, summary="Serve standalone SFTP Browser tool")
def read_sftp_tool():
    """Serve the standalone SFTP File Browser tool."""
    return _serve_html("sftp-browser.html")


@app.get("/cron", response_class=HTMLResponse, summary="Serve Cron Visualizer tool")
def read_cron_tool():
    """Serve the Cron Visualizer tool (Unix, Quartz, AWS EventBridge, GitHub Actions)."""
    return _serve_html("cron.html")


@app.get("/vault", response_class=HTMLResponse, summary="Serve Secret Vault tool")
def read_vault_tool():
    """Serve the Secret Vault tool (KeePass-style encrypted secrets manager)."""
    return _serve_html("vault.html")


@app.get("/db-manager", response_class=HTMLResponse, summary="Serve DevDB Manager tool")
def read_db_manager_tool():
    """Serve the DevDB Manager tool (unified encrypted database viewer and manager)."""
    return _serve_html("db-manager.html")


@app.get("/file-converter", response_class=HTMLResponse, summary="Serve File Format Converter tool")
def read_file_converter_tool():
    """Serve the File Format Converter tool.

    Supports: JSON, CSV, YAML, XML, XLSX, Markdown, Images, PDF, DOCX.
    """
    return _serve_html("file-converter.html")


def _conv_xlsx_to_csv_json(content: bytes, target_format: str) -> Response:
    """Convert XLSX bytes to CSV or JSON and return the appropriate Response."""
    import io  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    try:
        import openpyxl  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError as e:
        raise HTTPException(status_code=503, detail=_OPENPYXL_MISSING) from e
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Spreadsheet is empty")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [
        dict(zip(headers, [str(v) if v is not None else "" for v in row]))
        for row in rows[1:]
    ]
    if target_format == "csv":
        import csv as csv_mod  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
        buf = io.StringIO()
        writer = csv_mod.DictWriter(buf, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": 'attachment; filename="converted.csv"'})
    return Response(content=json.dumps(data, indent=2), media_type="application/json",
                    headers={"Content-Disposition": 'attachment; filename="converted.json"'})


def _conv_csv_to_xlsx(content: bytes) -> Response:
    """Convert CSV bytes to XLSX and return the appropriate Response."""
    import io  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    import csv as csv_mod  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    try:
        import openpyxl  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError as e:
        raise HTTPException(status_code=503, detail=_OPENPYXL_MISSING) from e
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv_mod.reader(io.StringIO(text))
    rows = list(reader)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="converted.xlsx"'},
    )


def _conv_json_to_xlsx(content: bytes) -> Response:
    """Convert JSON bytes to XLSX and return the appropriate Response."""
    import io  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    try:
        import openpyxl  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError as e:
        raise HTTPException(status_code=503, detail=_OPENPYXL_MISSING) from e
    data = json.loads(content)
    if not isinstance(data, list):
        data = [data]
    wb = openpyxl.Workbook()
    ws = wb.active
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([str(row.get(h, "")) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="converted.xlsx"'},
    )


def _conv_pdf_to_txt(content: bytes) -> Response:
    """Extract text from PDF bytes and return a plain-text Response."""
    import io  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    try:
        import pypdf  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError as e:
        raise HTTPException(status_code=503, detail="pypdf is not installed. Run: pip install pypdf") from e
    reader = pypdf.PdfReader(io.BytesIO(content))
    pages_text = [page.extract_text() or "" for page in reader.pages]
    full_text = "\n\n".join(pages_text)
    return Response(content=full_text, media_type="text/plain",
                    headers={"Content-Disposition": 'attachment; filename="converted.txt"'})


def _conv_docx_to_txt(content: bytes) -> Response:
    """Extract text from DOCX bytes and return a plain-text Response."""
    import io  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    try:
        import docx  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError as e:
        raise HTTPException(status_code=503, detail="python-docx is not installed. Run: pip install python-docx") from e
    doc = docx.Document(io.BytesIO(content))
    text = "\n".join(para.text for para in doc.paragraphs)
    return Response(content=text, media_type="text/plain",
                    headers={"Content-Disposition": 'attachment; filename="converted.txt"'})


def _source_to_html(src_ext: str, content: bytes) -> str:
    """Convert document bytes to a raw HTML string for PDF rendering."""
    import io  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    if src_ext in ("docx", "doc"):
        try:
            import mammoth  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
        except ImportError as e:
            raise HTTPException(status_code=503, detail="mammoth is not installed. Run: pip install mammoth") from e
        return mammoth.convert_to_html(io.BytesIO(content)).value
    if src_ext in ("md", "markdown"):
        import html as _html_mod  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
        try:
            import markdown as md_lib  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
            return md_lib.markdown(
                content.decode("utf-8", errors="replace"),
                extensions=["tables", "fenced_code"],
            )
        except ImportError:
            return "<pre>" + _html_mod.escape(content.decode("utf-8", errors="replace")) + "</pre>"
    return content.decode("utf-8", errors="replace")  # html / htm


def _conv_any_to_pdf(src_ext: str, content: bytes) -> Response:
    """Convert DOCX/DOC/HTML/MD/Markdown bytes to PDF using weasyprint."""
    try:
        import weasyprint  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError as e:
        raise HTTPException(status_code=503, detail="weasyprint is not installed. Run: pip install weasyprint") from e

    raw_html = _source_to_html(src_ext, content)

    # Step 2: Wrap in a full HTML document with print-friendly CSS
    full_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Georgia, 'Times New Roman', serif; font-size: 12pt; line-height: 1.6;
          margin: 2cm; color: #111; }}
  h1,h2,h3,h4,h5,h6 {{ font-family: Arial, Helvetica, sans-serif; margin-top: 1.2em; }}
  h1 {{ font-size: 22pt; }} h2 {{ font-size: 17pt; }} h3 {{ font-size: 14pt; }}
  p {{ margin: 0.5em 0 0.8em; }}
  table {{ border-collapse: collapse; width: 100%; margin: 1em 0; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; text-align: left; }}
  th {{ background: #f0f0f0; font-weight: bold; }}
  pre, code {{ font-family: 'Courier New', monospace; font-size: 10pt;
               background: #f5f5f5; padding: 0.2em 0.4em; border-radius: 3px; }}
  pre {{ padding: 0.8em; white-space: pre-wrap; word-break: break-word; }}
  img {{ max-width: 100%; height: auto; }}
  a {{ color: #1a56db; }}
  @page {{ margin: 2cm; }}
</style>
</head>
<body>{raw_html}</body>
</html>"""

    # Step 3: Render to PDF with a safe URL fetcher to prevent SSRF/LFI
    try:
        from weasyprint import default_url_fetcher as _default_url_fetcher  # noqa: PLC0415  # pylint: disable=import-outside-toplevel
    except ImportError:
        _default_url_fetcher = None

    def _safe_url_fetcher(url: str):
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in ('', 'data'):
            raise ValueError(
                f"Blocked disallowed URL scheme '{parsed.scheme}' in PDF conversion: {url!r}"
            )
        if _default_url_fetcher is not None:
            return _default_url_fetcher(url)
        raise ValueError(f"No URL fetcher available for: {url!r}")

    pdf_bytes = weasyprint.HTML(string=full_html, url_fetcher=_safe_url_fetcher).write_pdf()
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": 'attachment; filename="converted.pdf"'})


def _check_content_length_header(cl: str | None) -> None:
    """Raise HTTPException 400/413 if the Content-Length header is invalid or too large."""
    if not cl:
        return
    try:
        cl_int = int(cl)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid Content-Length header") from exc
    if cl_int < 0:
        raise HTTPException(status_code=400, detail="Invalid Content-Length header")
    if cl_int > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Upload too large (limit {MAX_UPLOAD_SIZE // 1024 // 1024} MB)",
        )


@app.post(
    "/api/convert",
    summary="Convert a file from one format to another (server-side)",
    responses={
        400: {"description": "Invalid content-length or unsupported conversion"},
        413: {"description": "Upload too large"},
        503: {"description": "Required conversion library not installed"},
    },
)
async def convert_file(
    request: Request,
    file: Annotated[UploadFile, File(...)],
    target_format: Annotated[str, Form(...)],
):
    """
    Server-side file format conversion endpoint.

    Handles conversions that require Python libraries:
    - XLSX ↔ CSV / JSON
    - PDF → TXT
    - DOCX / DOC → PDF (requires weasyprint)
    - Markdown → PDF (requires weasyprint)
    - HTML → PDF (requires weasyprint)
    """
    target_format = target_format.lower().strip()
    original_name = (file.filename or "file").lower()
    src_ext = original_name.rsplit(".", 1)[-1] if "." in original_name else ""

    _check_content_length_header(request.headers.get("content-length"))
    content = await file.read(MAX_UPLOAD_SIZE + 1)
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Upload too large (limit {MAX_UPLOAD_SIZE // 1024 // 1024} MB)",
        )

    if src_ext == "xlsx" and target_format in ("csv", "json"):
        return _conv_xlsx_to_csv_json(content, target_format)
    if src_ext == "csv" and target_format == "xlsx":
        return _conv_csv_to_xlsx(content)
    if src_ext == "json" and target_format == "xlsx":
        return _conv_json_to_xlsx(content)
    if src_ext == "pdf" and target_format == "txt":
        return _conv_pdf_to_txt(content)
    if src_ext == "docx" and target_format == "txt":
        return _conv_docx_to_txt(content)
    if target_format == "pdf" and src_ext in ("docx", "doc", "html", "htm", "md", "markdown"):
        return _conv_any_to_pdf(src_ext, content)
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported server-side conversion: {src_ext} → {target_format}"
    )


@app.get(
    "/api/vault",
    summary="Get encrypted vault blob",
    responses={401: {"description": "Session token missing or expired"}},
)
def get_vault(request: Request):
    """Return the raw encrypted vault blob from the DevDB 'vault' store.
    Backward-compatible shim — the server never decrypts vault contents.
    """
    require_unlocked(request)
    store = _db.get_store("vault")
    return store if store else {"encrypted_blob": ""}


@app.post(
    "/api/vault",
    summary="Save encrypted vault blob",
    responses={
        401: {"description": "Session token missing or expired"},
        500: {"description": "Failed to save vault"},
    },
)
def save_vault(data: dict, request: Request):
    """Persist the encrypted vault blob into the DevDB 'vault' store.
    Backward-compatible shim — the server never decrypts vault contents.
    """
    require_unlocked(request)
    try:
        _db.set_store("vault", data)
        _db.save()
        return {"status": "ok"}
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to save vault: %s", e)
        raise HTTPException(status_code=500, detail="Failed to save vault") from e


@app.post(
    "/api/shorten",
    summary="Create a short URL",
    responses={
        400: {"description": "Invalid or empty URL"},
        500: {"description": "Failed to generate unique short ID"},
    },
)
def shorten_url(req: ShortenRequest, request: Request):
    """
    Create and store a 6-character short identifier for the provided URL and return
    the short link and original URL.

    The input URL is trimmed of surrounding whitespace; if it lacks an HTTP scheme,
    ``https://`` is prepended. A random 6-character alphanumeric ``short_id`` is
    generated, stored in the in-memory datastore, and used to build the redirectable
    short URL from ``request.base_url``.

    Parameters:
        req (ShortenRequest): Request model containing the `url` to shorten.
        request (Request): FastAPI request used to derive the application's base URL.

    Returns:
        dict: {
            "short_id": str — the generated 6-character identifier,
            "short_url": str — the full short URL that redirects to the original,
            "original_url": str — the normalized original URL stored for this id
        }
    """
    url = req.url.strip()
    if not url:
        raise HTTPException(status_code=400, detail="URL cannot be empty")

    parsed = urllib.parse.urlparse(url)
    if not parsed.scheme:
        url = "https://" + url
        parsed = urllib.parse.urlparse(url)
    elif parsed.scheme not in ("http", "https"):
        raise HTTPException(status_code=400, detail="Unsupported URL scheme")

    if not parsed.scheme or not parsed.netloc:
        raise HTTPException(status_code=400, detail="Invalid URL format")

    alphabet = string.ascii_letters + string.digits
    for _ in range(10):
        short_id = "".join(secrets.choice(alphabet) for _ in range(6))
        if short_id not in url_db:
            break
    else:
        raise HTTPException(status_code=500, detail="Failed to generate unique short ID")

    # Store in memory and persist to DevDB
    url_db[short_id] = url
    _save_url_db(url_db)

    # Return the full short URL
    base_url = str(request.base_url).rstrip("/")
    short_url = f"{base_url}/r/{short_id}"
    return {"short_id": short_id, "short_url": short_url, "original_url": url}


@app.get(
    "/r/{short_id}",
    summary="Redirect to original URL",
    responses={404: {"description": "Short URL not found"}},
)
def redirect_short_url(short_id: str):
    """
    Redirects a short identifier to the stored original URL.

    Returns:
        RedirectResponse: A 302 redirect response to the original URL.

    Raises:
        HTTPException: If the provided `short_id` does not exist (status code 404).
    """
    if short_id in url_db:
        return RedirectResponse(url=url_db[short_id], status_code=302)
    raise HTTPException(status_code=404, detail="Short URL not found.")


async def _read_upload_stream(file: UploadFile, max_size: int) -> tuple[bytearray, bool]:
    """Read file in 1 MB chunks; detect null bytes in the first chunk; enforce size limit."""
    raw_bytes = bytearray()
    null_detected = False
    chunk_size = 1024 * 1024
    while chunk := await file.read(chunk_size):
        if not null_detected and not raw_bytes and b"\x00" in chunk[:512]:
            null_detected = True
        raw_bytes.extend(chunk)
        if len(raw_bytes) > max_size:
            raise HTTPException(status_code=413, detail="File too large. Exceeds 50MB limit.")
    return raw_bytes, null_detected


@app.post(
    "/upload",
    summary="Upload a text file for diffing",
    responses={
        400: {"description": "Binary file or invalid content type"},
        413: {"description": "File too large (50 MB limit)"},
        500: {"description": "Server error processing file"},
    },
)
async def upload_file(file: Annotated[UploadFile, File(...)]):
    """Accept an uploaded text file and return its content and metadata.

    Raises 400 for binary files, 413 if over 50 MB, 500 on unexpected errors.
    """
    binary_mimes = ("image/", "video/", "audio/", "application/pdf",
                    "application/zip", _MIME_OCTET_STREAM)
    if file.content_type and any(file.content_type.startswith(b) for b in binary_mimes):
        raise HTTPException(
            status_code=400,
            detail=f"Only text-based files are supported. Received: {file.content_type}"
        )
    try:
        raw_bytes, null_detected = await _read_upload_stream(file, 50 * 1024 * 1024)
        if null_detected:
            raise HTTPException(
                status_code=400,
                detail=f'"{file.filename}" appears to be a binary file and cannot be diffed.'
            )
        content = raw_bytes.decode("utf-8", errors="replace")
        return {"filename": file.filename, "content": content, "size_bytes": len(raw_bytes)}
    except HTTPException:
        raise
    except Exception as e:  # pylint: disable=broad-exception-caught
        raise HTTPException(status_code=500, detail="Server error processing file") from e


@app.get(
    "/api/collections",
    summary="Get API Tester Collections",
    responses={500: {"description": "Failed to read collections"}},
)
def get_collections():
    """Read saved collections from the DevDB 'collections' store.
    Backward-compatible shim for api-tester.js.
    """
    store = _db.get_store("collections")
    return store if store else {"items": []}


@app.post(
    "/api/collections",
    summary="Save API Tester Collections",
    responses={500: {"description": "Failed to save collections"}},
)
def save_collections(data: dict):
    """Persist collections into the DevDB 'collections' store.
    Backward-compatible shim for api-tester.js.
    """
    try:
        _db.set_store("collections", data)
        _db.save()
        return {"status": "ok"}
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to save collections: %s", e)
        raise HTTPException(status_code=500, detail="Failed to save collections") from e


class ProxyRequest(BaseModel):
    """Request body for the /api/proxy endpoint."""

    url: str
    method: str = "GET"
    headers: dict = {}
    body: str | None = None


def _check_ip_not_private(ip_str: str) -> None:
    """Raise HTTPException 403 if the IP is private/reserved/loopback."""
    try:
        ip_obj = ipaddress.ip_address(ip_str)
        if (ip_obj.is_loopback or ip_obj.is_private or ip_obj.is_link_local
                or ip_obj.is_multicast or ip_obj.is_reserved):
            raise HTTPException(
                status_code=403,
                detail=f"Access to private/reserved IP addresses is forbidden: {ip_str}",
            )
        if ip_str.startswith("169.254."):
            raise HTTPException(status_code=403, detail="Access to cloud metadata endpoints is forbidden")
    except ValueError:
        pass  # intentionally ignored: non-IP strings (hostnames) are not checked


_HOP_BY_HOP_HEADERS = frozenset(("host", "connection", "origin", "referer", "accept-encoding"))


def _filter_proxy_headers(headers: dict) -> dict:
    return {k: v for k, v in headers.items() if k.lower() not in _HOP_BY_HOP_HEADERS}


def _execute_proxy_request(request_obj) -> dict:
    """Run a urllib request and return a normalised proxy-response dict."""
    try:
        with urllib.request.urlopen(request_obj, timeout=15) as resp:  # nosec B310
            return {
                "proxy_response": True,
                "status": resp.status,
                "headers": dict(resp.headers),
                "body": resp.read().decode("utf-8", errors="replace"),
            }
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else ""
        except (OSError, ValueError):
            body = ""
        return {"proxy_response": True, "status": e.code, "headers": dict(e.headers), "body": body}


def _resolve_target_ips(hostname: str, port: int | None, scheme: str) -> None:
    """Resolve hostname to IP addresses and reject any private/reserved ones."""
    try:
        addr_info = socket.getaddrinfo(
            hostname,
            port or (443 if scheme == 'https' else 80),
            socket.AF_UNSPEC,
            socket.SOCK_STREAM,
        )
    except (socket.gaierror, socket.herror) as e:
        raise HTTPException(status_code=400, detail=f"DNS resolution failed: {e}") from e
    for _, _, _, _, sockaddr in addr_info:
        _check_ip_not_private(sockaddr[0])


@app.post(
    "/api/proxy",
    summary="Bypass CORS for API Tester",
    responses={
        400: {"description": "Invalid URL or DNS failure"},
        403: {"description": "Target IP is private or reserved"},
        500: {"description": "Proxy request failed"},
    },
)
async def proxy_request(req: ProxyRequest):
    """Provides a local CORS bypass proxy using urllib for the API tester tool."""
    try:
        parsed = urllib.parse.urlparse(req.url)
        if parsed.scheme not in ('http', 'https'):
            raise HTTPException(status_code=400, detail="Only HTTP and HTTPS schemes are allowed")
        if not parsed.hostname:
            raise HTTPException(status_code=400, detail="Invalid URL: no hostname")
        if parsed.hostname not in ALLOWED_PROXY_HOSTS:
            raise HTTPException(status_code=400, detail="Target host is not allowed")

        _resolve_target_ips(parsed.hostname, parsed.port, parsed.scheme)

        req_body = req.body.encode('utf-8') if req.body else None
        # Reconstruct the URL using the validated components to clear CodeQL dataflow taint.
        safe_host = next(h for h in ALLOWED_PROXY_HOSTS if h == parsed.hostname)
        safe_netloc = f"{safe_host}:{parsed.port}" if parsed.port else safe_host
        safe_url = urllib.parse.urlunparse((
            parsed.scheme, safe_netloc, parsed.path, parsed.params, parsed.query, parsed.fragment
        ))
        request_obj = urllib.request.Request(
            safe_url, data=req_body,
            headers=_filter_proxy_headers(req.headers),
            method=req.method.upper(),
        )
        return _execute_proxy_request(request_obj)
    except HTTPException:
        raise
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Proxy request failed: %s", e)
        raise HTTPException(status_code=500, detail="Proxy request failed") from e


@app.get(
    "/api/ssh/profiles",
    summary="Get SSH Profiles",
    responses={401: {"description": "Session token missing or expired"}},
)
def get_ssh_profiles(request: Request):
    """Return the encrypted SSH profiles blob from the DevDB 'ssh_profiles' store.
    Backward-compatible shim — server never decrypts profile contents.
    """
    require_unlocked(request)
    store = _db.get_store("ssh_profiles")
    return store if store else {"encrypted_blob": ""}


@app.post(
    "/api/ssh/profiles",
    summary="Save SSH Profiles",
    responses={
        401: {"description": "Session token missing or expired"},
        500: {"description": "Failed to save SSH profiles"},
    },
)
def save_ssh_profiles(data: dict, request: Request):
    """Persist the encrypted SSH profiles blob into the DevDB 'ssh_profiles' store.
    Backward-compatible shim — server never decrypts profile contents.
    """
    require_unlocked(request)
    try:
        _db.set_store("ssh_profiles", data)
        _db.save()
        return {"status": "ok"}
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to save SSH profiles: %s", e)
        raise HTTPException(status_code=500, detail="Failed to save SSH profiles") from e


# ─── Server-side session store ───────────────────────────────────────────────
# Tokens are issued by /api/auth/session after the client verifies the
# master key, and are required by all DevDB endpoints.
_sessions: dict[str, float] = {}   # token → unix expiry
_SESSION_TTL = 8 * 3600            # 8 hours (matches auth-guard.js SESSION_MS)


def require_unlocked(request: Request) -> None:
    """Raise 401 if the request does not carry a valid server-side session token."""
    token = request.headers.get("X-Session-Token", "").strip()
    if not token:
        raise HTTPException(
            status_code=401,
            detail="Session token required. Call POST /api/auth/session first.",
        )
    expiry = _sessions.get(token)
    if expiry is None or time.time() > expiry:
        _sessions.pop(token, None)
        raise HTTPException(status_code=401, detail="Session expired or invalid.")


# ─── DevDB Unified API ────────────────────────────────────────────────────────
# These endpoints expose the DevDB engine directly for the DB Manager UI
# and any future tools that want to read/write named stores.

_ALLOWED_STORES = {"vault", "collections", "ssh_profiles", "url_db", "app_prefs"}

# Only allow printable, non-shell-special characters for WSL distro names.
_DISTRO_NAME_RE = re.compile(r'^[A-Za-z0-9_.\-]+$')


@app.get(
    "/api/db/meta",
    summary="Get DevDB metadata",
    responses={401: {"description": "Session token missing or expired"}},
)
def db_meta(request: Request):
    """Return database metadata: path, file size, stores list, encryption status."""
    require_unlocked(request)
    m = _db.meta()
    return {
        "path":      str(_DB_PATH),
        "size":      _db.file_size(),
        "encrypted": _db.is_encrypted(),
        "stores":    _db.store_sizes(),
        "meta":      m,
    }


@app.get(
    "/api/db/store/{name}",
    summary="Read a named DevDB store",
    responses={
        400: {"description": "Unknown store name"},
        401: {"description": "Session token missing or expired"},
    },
)
def db_get_store(name: str, request: Request):
    """Return the raw contents of the named store.  Restricted to known store names."""
    require_unlocked(request)
    if name not in _ALLOWED_STORES:
        raise HTTPException(status_code=400, detail=f"Unknown store: {name!r}")
    return _db.get_store(name)


@app.post(
    "/api/db/store/{name}",
    summary="Write a named DevDB store",
    responses={
        400: {"description": "Unknown store name"},
        401: {"description": "Session token missing or expired"},
        500: {"description": "Failed to write store"},
    },
)
def db_set_store(name: str, data: dict, request: Request):
    """Replace the named store with the supplied data and flush to disk."""
    require_unlocked(request)
    if name not in _ALLOWED_STORES:
        raise HTTPException(status_code=400, detail=f"Unknown store: {name!r}")
    try:
        _db.set_store(name, data)
        _db.save()
        if name == "url_db":
            url_db.clear()
            url_db.update(_db.get_store("url_db") or {})
        return {"status": "ok", "store": name}
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to write store %r: %s", name, e)
        raise HTTPException(status_code=500, detail="Failed to write store") from e


@app.get(
    "/api/db/export",
    summary="Export full DevDB as a .dsb file",
    responses={
        401: {"description": "Session token missing or expired"},
        500: {"description": "Failed to export database"},
    },
)
def db_export(request: Request):
    """Stream the raw .dsb binary as a file download."""
    require_unlocked(request)
    try:
        raw = _db.export_bytes()
        return Response(
            content=raw,
            media_type=_MIME_OCTET_STREAM,
            headers={"Content-Disposition": 'attachment; filename="devdb.dsb"'},
        )
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to export DevDB: %s", e)
        raise HTTPException(status_code=500, detail="Failed to export database") from e


@app.post(
    "/api/db/import",
    summary="Import a .dsb file into DevDB",
    responses={
        400: {"description": "Invalid .dsb format"},
        401: {"description": "Session token missing or expired"},
        413: {"description": "Import file too large (50 MB limit)"},
        500: {"description": "Failed to import database"},
    },
)
async def db_import(request: Request, file: Annotated[UploadFile, File(...)]):
    """Accept a .dsb upload and merge its stores into the running DevDB."""
    require_unlocked(request)
    max_import_size = 50 * 1024 * 1024  # 50 MB
    try:
        raw = await file.read(max_import_size + 1)
        if len(raw) > max_import_size:
            raise HTTPException(status_code=413, detail="Import file too large (50 MB limit)")
        imported = DevDB.from_bytes(raw)  # parses & validates the binary format
        # Merge all stores from the imported file (skip unknown store names)
        url_db_updated = False
        for store_name in imported.list_stores():
            if store_name in _ALLOWED_STORES:
                _db.set_store(store_name, imported.get_store(store_name))
                if store_name == "url_db":
                    url_db_updated = True
        _db.save()
        if url_db_updated:
            url_db.clear()
            url_db.update(_db.get_store("url_db") or {})
        return {"status": "ok", "imported_stores": imported.list_stores()}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to import DevDB: %s", e)
        raise HTTPException(status_code=500, detail="Failed to import database") from e

# ─── Auth — Master Password Management ──────────────────────────────────────
# Client-side password verification: server stores a challenge blob (AES-encrypted
# known plaintext).  The plaintext password never leaves the browser.


@app.get(
    "/api/auth/status",
    summary="Check if master password is configured",
    responses={},
)
def auth_status():
    """Return whether the master encryption password has been set up."""
    prefs = _db.get_store("app_prefs") or {}
    vault = _db.get_store("vault") or {}
    return {
        "is_setup":       bool(prefs.get("master_setup_done")),
        "vault_has_data": bool(vault.get("salt")),
    }


@app.get(
    "/api/auth/challenge",
    summary="Get password verification challenge",
    responses={404: {"description": "Master password not configured"}},
)
def auth_challenge():
    """Return the stored salt + encrypted-verify-blob for client-side password checking."""
    prefs = _db.get_store("app_prefs") or {}
    if not prefs.get("master_setup_done"):
        raise HTTPException(status_code=404, detail="Master password not configured")
    return {
        "salt":        prefs.get("master_salt", ""),
        "verify_blob": prefs.get("master_verify_blob", ""),
        "verify_iv":   prefs.get("master_verify_iv", ""),
    }


@app.post(
    "/api/auth/setup",
    summary="Initialize master password (first-time setup)",
    responses={
        400: {"description": "Missing required fields"},
        409: {"description": "Master password already configured"},
    },
)
def auth_setup(data: dict):
    """One-time setup: store the PBKDF2 salt and AES verification blob in app_prefs.
    Called by vault.js after the user sets their master password for the first time.
    Expects: {salt, verify_blob, verify_iv}
    """
    prefs = _db.get_store("app_prefs") or {}
    if prefs.get("master_setup_done"):
        raise HTTPException(status_code=409, detail="Master password already configured")

    salt        = str(data.get("salt",        "")).strip()
    verify_blob = str(data.get("verify_blob", "")).strip()
    verify_iv   = str(data.get("verify_iv",   "")).strip()

    if not salt or not verify_blob or not verify_iv:
        raise HTTPException(
            status_code=400,
            detail="Missing required fields: salt, verify_blob, verify_iv",
        )

    prefs.update({
        "master_setup_done":  True,
        "master_salt":        salt,
        "master_verify_blob": verify_blob,
        "master_verify_iv":   verify_iv,
    })
    _db.set_store("app_prefs", prefs)
    _db.save()
    return {"status": "ok"}


@app.post(
    "/api/auth/update-challenge",
    summary="Update master password challenge after password change",
    responses={
        400: {"description": "Missing required fields"},
        401: {"description": "Session token missing or expired"},
        404: {"description": "Master password not yet configured"},
    },
)
def auth_update_challenge(data: dict, request: Request):
    """Replace the verification challenge when the master password is changed.
    Expects: {salt, verify_blob, verify_iv}
    """
    require_unlocked(request)
    prefs = _db.get_store("app_prefs") or {}
    if not prefs.get("master_setup_done"):
        raise HTTPException(status_code=404, detail="Master password not yet configured")

    salt        = str(data.get("salt",        "")).strip()
    verify_blob = str(data.get("verify_blob", "")).strip()
    verify_iv   = str(data.get("verify_iv",   "")).strip()

    if not salt or not verify_blob or not verify_iv:
        raise HTTPException(
            status_code=400,
            detail="Missing required fields: salt, verify_blob, verify_iv",
        )

    prefs.update({
        "master_salt":        salt,
        "master_verify_blob": verify_blob,
        "master_verify_iv":   verify_iv,
    })
    _db.set_store("app_prefs", prefs)
    _db.save()
    # Revoke all existing session tokens so old sessions cannot continue
    # after a master password rotation.
    _sessions.clear()
    return {"status": "ok"}


@app.post(
    "/api/auth/session",
    summary="Exchange verified master key for a server-side session token",
    responses={
        400: {"description": "Missing key_hex"},
        401: {"description": "Invalid master key or key verification failed"},
        404: {"description": "Master password not configured"},
    },
)
def auth_session(data: dict):  # pylint: disable=too-many-locals
    """Verify the PBKDF2-derived key (hex) against the stored challenge and issue a session token.

    The client sends {key_hex: <hex>} where key_hex is the AES key derived from
    the master password using PBKDF2-SHA1(50 000 iter, 32-byte output).  The server
    decrypts the stored verify_blob to confirm key correctness without ever seeing
    the plaintext password.  On success a session token valid for 8 hours is returned.
    """
    prefs = _db.get_store("app_prefs") or {}
    if not prefs.get("master_setup_done"):
        raise HTTPException(status_code=404, detail="Master password not configured")

    key_hex = str(data.get("key_hex", "")).strip()
    if not key_hex:
        raise HTTPException(status_code=400, detail="Missing key_hex")

    try:
        import base64 as _b64  # pylint: disable=import-outside-toplevel
        from cryptography.hazmat.primitives.ciphers import (  # pylint: disable=import-outside-toplevel
            Cipher,
            algorithms,
            modes,
        )
        from cryptography.hazmat.backends import default_backend  # pylint: disable=import-outside-toplevel

        key        = bytes.fromhex(key_hex)
        verify_iv  = bytes.fromhex(prefs["master_verify_iv"])
        ciphertext = _b64.b64decode(prefs["master_verify_blob"])

        cipher = Cipher(algorithms.AES(key), modes.CBC(verify_iv), backend=default_backend())  # NOSONAR — CBC required for CryptoJS client compatibility; verify_blob is a one-way challenge, not sensitive data
        decryptor = cipher.decryptor()
        padded = decryptor.update(ciphertext) + decryptor.finalize()
        # Remove PKCS7 padding
        pad_len = padded[-1]
        if pad_len < 1 or pad_len > 16:
            raise ValueError("Invalid PKCS7 padding")
        plaintext = padded[:-pad_len].decode("utf-8", errors="strict")

        if plaintext != "DEVSUITE_MASTER_OK":
            raise HTTPException(status_code=401, detail="Invalid master key")
    except HTTPException:
        raise
    except Exception as exc:  # pylint: disable=broad-exception-caught
        raise HTTPException(status_code=401, detail="Key verification failed") from exc

    token = secrets.token_urlsafe(32)
    _sessions[token] = time.time() + _SESSION_TTL
    return {"session_token": token, "expires_in": _SESSION_TTL}


def _create_known_hosts(path: str) -> None:
    """Create an empty known_hosts file with mode 600."""
    with open(path, "w", encoding="utf-8") as _fh:
        pass  # intentionally empty: creates the file with no initial content
    os.chmod(path, 0o600)


def _append_known_hosts(path: str, data: bytes) -> None:
    """Append a host-key entry to the known_hosts file."""
    with open(path, "ab") as fh:
        fh.write(data)


async def _ssh_keyscan(host: str, port: int) -> bytes:
    """Fetch the public key blob for host:port via ssh-keyscan."""
    proc = await asyncio.create_subprocess_exec(
        "ssh-keyscan", "-p", str(port), "-H", host,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    try:
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=15)
    except asyncio.TimeoutError as exc:
        raise RuntimeError(f"ssh-keyscan timed out for {host}:{port}") from exc
    if proc.returncode != 0 or not stdout.strip():
        raise RuntimeError(
            f"Could not retrieve host key for {host}:{port}. Is the host reachable?"
        )
    return stdout


async def _ssh_key_fingerprint(key_data: bytes, host: str, port: int) -> str:
    """Return the SHA-256 (or MD5) fingerprint string for a raw ssh-keyscan blob."""
    keygen = await asyncio.create_subprocess_exec(
        "ssh-keygen", "-l", "-f", "-",
        stdin=asyncio.subprocess.PIPE,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        kg_out, _ = await asyncio.wait_for(keygen.communicate(input=key_data), timeout=10)
    except asyncio.TimeoutError as exc:
        raise RuntimeError(f"ssh-keygen fingerprint timed out for {host}:{port}") from exc
    for token in kg_out.decode(errors="replace").split():
        if token.startswith("SHA256:") or token.startswith("MD5:"):
            return token
    return ""


async def _ensure_host_key(
    host: str,
    port: int,
    approve_host=None,
) -> str:
    """Ensure ~/.ssh/known_hosts exists and contains a pinned entry for host:port.

    For an unknown host: fetches the key via ssh-keyscan, computes its fingerprint,
    then calls ``await approve_host(host, port, fingerprint, key_line)`` if provided.
    Raises HostKeyApprovalRequired when no callback is supplied.
    Returns the path to the known_hosts file.
    """
    known_hosts_path = os.path.expanduser("~/.ssh/known_hosts")
    ssh_dir = os.path.dirname(known_hosts_path)

    os.makedirs(ssh_dir, mode=0o700, exist_ok=True)
    if not os.path.exists(known_hosts_path):
        await asyncio.to_thread(_create_known_hosts, known_hosts_path)

    lookup = f"[{host}]:{port}" if port != 22 else host
    check = await asyncio.create_subprocess_exec(
        "ssh-keygen", "-F", lookup, "-f", known_hosts_path,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    await check.wait()
    if check.returncode == 0:
        return known_hosts_path

    key_data = await _ssh_keyscan(host, port)
    fingerprint = await _ssh_key_fingerprint(key_data, host, port)

    if approve_host is None:
        raise HostKeyApprovalRequired(host, port, fingerprint)

    approved = await approve_host(host, port, fingerprint, key_data)
    if not approved:
        raise RuntimeError(
            f"Host key for {host}:{port} (fingerprint {fingerprint}) was rejected by the user."
        )

    await asyncio.to_thread(_append_known_hosts, known_hosts_path, key_data)
    return known_hosts_path


class HostKeyApprovalRequired(Exception):
    """Raised by _ensure_host_key when no approve_host callback is provided."""
    def __init__(self, host: str, port: int, fingerprint: str):
        super().__init__(f"Host key approval required for {host}:{port}")
        self.host = host
        self.port = port
        self.fingerprint = fingerprint


async def _ws_check_origin(websocket: WebSocket) -> bool:
    """Return True if origin is valid; close the socket and return False otherwise."""
    origin = websocket.headers.get("origin", "")
    host = websocket.headers.get("host", "")
    if not origin:
        await websocket.close(code=1008, reason=_ERR_ORIGIN_REQUIRED)
        return False
    if origin not in _ALLOWED_ORIGINS and host and not origin.endswith(f"//{host}"):
        await websocket.close(code=1008, reason=_ERR_ORIGIN_NOT_ALLOWED)
        return False
    return True


def _build_ssh_connect_kwargs(
    host: str,
    port: int,
    username: str,
    password: str | None,
    private_key: str | None,
    known_hosts_path: str,
) -> dict:
    """Build the keyword arguments dict for asyncssh.connect()."""
    kwargs: dict = {"host": host, "port": port, "username": username, "known_hosts": known_hosts_path}
    if password:
        kwargs["password"] = password
    if private_key:
        kwargs["client_keys"] = [asyncssh.import_private_key(private_key)]
    return kwargs


async def _run_metrics_loop(websocket: WebSocket, conn) -> None:
    """Poll SSH server metrics every 2 s and push them over *websocket*."""
    script = (
        "cat /proc/stat | head -n 1; echo '---'; "
        "cat /proc/meminfo | grep -E '^(MemTotal|MemAvailable|SwapTotal|SwapFree):'; echo '---'; "
        "df -m | awk 'NR>1 {print $1, $2, $3, $4, $5, $6}'; echo '---'; "
        "cat /proc/uptime | awk '{print $1}'"
    )
    await websocket.send_json({"status": "connected"})
    prev_idle = 0
    prev_total = 0
    while True:
        try:
            res = await asyncio.wait_for(conn.run(script), timeout=5.0)
            if res.exit_status == 0:
                parts = res.stdout.strip().split('---')
                if len(parts) == 4:
                    payload, prev_idle, prev_total = _parse_ssh_metrics(parts, prev_idle, prev_total)
                    await websocket.send_json(payload)
            await asyncio.sleep(2)
        except asyncio.TimeoutError:
            continue
        except Exception as e:  # pylint: disable=broad-exception-caught
            logger.debug("ssh_dashboard iteration error: %s", e)
            break


def _try_resize_ssh_process(process, data: str) -> bool:
    """Apply terminal resize if *data* is a resize escape; return True if handled."""
    if not data.startswith("\x1b[resize;"):
        return False
    parts = data.split(";")
    if len(parts) == 3:
        try:
            cols, rows = int(parts[1]), int(parts[2].strip("m"))
            process.change_terminal_size(cols, rows, 0, 0)
        except Exception:  # pylint: disable=broad-exception-caught
            logger.debug("Terminal resize failed (ignored)", exc_info=True)
    return True


async def _run_ssh_terminal_session(websocket: WebSocket, conn) -> None:
    """Run the interactive SSH terminal I/O loop over an established connection."""
    async with conn.create_process(term_type='xterm-256color') as process:

        async def read_from_ssh():
            try:
                while data := await process.stdout.read(4096):
                    await websocket.send_text(str(data))
            except Exception:  # pylint: disable=broad-exception-caught
                logger.debug("read_from_ssh: stream ended or error", exc_info=True)

        async def write_to_ssh():
            try:
                while True:
                    data = await websocket.receive_text()
                    if _try_resize_ssh_process(process, data):
                        continue
                    process.stdin.write(data)
            except WebSocketDisconnect:
                process.terminate()
            except Exception:  # pylint: disable=broad-exception-caught
                logger.debug("write_to_ssh: error writing to SSH process", exc_info=True)

        await asyncio.gather(read_from_ssh(), write_to_ssh())


async def _ws_wait_for_host_key_response(websocket: WebSocket) -> bool:
    """Wait for a host_key_response message over *websocket*.

    Returns True if the user approved, False on timeout or rejection.
    """
    while True:
        try:
            raw = await websocket.receive_text()
        except asyncio.TimeoutError:
            return False
        try:
            msg = json.loads(raw)
            if msg.get("type") == "host_key_response":
                return bool(msg.get("approve", False))
        except (json.JSONDecodeError, AttributeError) as exc:
            logger.debug(
                "Ignored non-JSON message while waiting for"
                " host_key_response (len=%d) — %s",
                len(raw), exc,
            )


async def _terminal_ws_approve_host(
    websocket: WebSocket,
    h: str,
    p: int,
    fingerprint: str,
    _key_line: bytes,
) -> bool:
    """Send a host-key approval request over *websocket* and await the browser reply."""
    await websocket.send_json({
        "type": "host_key_approval",
        "host": h,
        "port": p,
        "fingerprint": fingerprint,
    })
    async with asyncio.timeout(60):
        return await _ws_wait_for_host_key_response(websocket)


@app.websocket("/api/ssh/terminal")
async def ssh_terminal(websocket: WebSocket):
    """WebSocket endpoint: interactive SSH terminal session."""
    if not await _ws_check_origin(websocket):
        return

    await websocket.accept()
    try:
        data = await websocket.receive_text()
        config = json.loads(data)
        ssh_host = config.get("host")
        port = int(config.get("port", 22))
        username = config.get("username")
        password = config.get("password")
        private_key = config.get("private_key")

        # Ensure known_hosts exists and has an entry for this host.
        # For new hosts the fingerprint is sent to the browser for user approval.
        await websocket.send_text(f"Verifying host key for {ssh_host}:{port}...\r\n")

        async def _ws_approve_host(h: str, p: int, fingerprint: str, _key_line: bytes) -> bool:
            return await _terminal_ws_approve_host(websocket, h, p, fingerprint, _key_line)

        try:
            known_hosts_path = await _ensure_host_key(ssh_host, port, approve_host=_ws_approve_host)
        except RuntimeError as exc:
            await websocket.send_text(f"\r\nHost key error: {exc}\r\n")
            await websocket.close()
            return

        connect_kwargs = _build_ssh_connect_kwargs(ssh_host, port, username, password, private_key, known_hosts_path)
        async with asyncssh.connect(**connect_kwargs) as conn:
            await _run_ssh_terminal_session(websocket, conn)
    except Exception as e:  # pylint: disable=broad-exception-caught
        try:
            await websocket.send_text(f"\r\nError: {e}\r\n")
            await websocket.close()
        except Exception:  # pylint: disable=broad-exception-caught
            logger.debug("ssh_terminal: failed to send error message to client", exc_info=True)


class SFTPRequest(BaseModel):
    """Request body for the /api/sftp/list endpoint."""

    host: str
    port: int = 22
    username: str
    password: str | None = None
    private_key: str | None = None
    path: str = "."
    approved_fingerprint: str | None = None


class SFTPDownloadRequest(BaseModel):
    """Request body for the /api/sftp/download endpoint."""

    host: str
    port: int = 22
    username: str
    password: str | None = None
    private_key: str | None = None
    path: str  # full remote file path
    approved_fingerprint: str | None = None


def _make_sftp_approve(approved_fingerprint: str | None):
    """Return an SFTP host-key approval callback that auto-approves a known fingerprint."""
    async def _approve(h: str, p: int, fingerprint: str, _key: bytes) -> bool:  # NOSONAR — asyncssh requires an awaitable callback
        if approved_fingerprint and approved_fingerprint == fingerprint:
            return True
        raise HTTPException(
            status_code=409,
            detail={
                "error": "host_key_approval_required",
                "host": h, "port": p, "fingerprint": fingerprint,
            },
        )
    return _approve


@app.post(
    "/api/sftp/list",
    summary="List files via SFTP",
    responses={
        409: {"description": "Host key approval required"},
        500: {"description": _ERR_SFTP_FAILED},
    },
)
async def sftp_list(req: SFTPRequest):
    """List files in a directory on a remote host via SFTP."""
    try:
        known_hosts_path = await _ensure_host_key(
            req.host, req.port, approve_host=_make_sftp_approve(req.approved_fingerprint)
        )
        connect_kwargs = _build_ssh_connect_kwargs(
            req.host, req.port, req.username, req.password, req.private_key, known_hosts_path
        )
        async with asyncssh.connect(**connect_kwargs) as conn:
            sftp = await conn.start_sftp_client()
            async with sftp:
                files = await sftp.readdir(req.path)
                result = []
                for f in files:
                    if f.filename in ('.', '..'):
                        continue
                    attrs = f.attrs
                    is_dir = stat.S_ISDIR(attrs.permissions) if attrs.permissions else False
                    result.append({
                        "name": f.filename,
                        "is_dir": is_dir,
                        "size": attrs.size,
                    })
                result.sort(key=lambda x: (not x['is_dir'], x['name'].lower()))
                return {"files": result, "cwd": req.path}
    except HTTPException:
        raise
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("SFTP list error: %s", e)
        raise HTTPException(status_code=500, detail=_ERR_SFTP_FAILED) from e


@app.post(
    "/api/sftp/download",
    summary="Download a file via SFTP",
    responses={
        409: {"description": "Host key approval required"},
        500: {"description": _ERR_SFTP_FAILED},
    },
)
async def sftp_download(req: SFTPDownloadRequest):
    """Stream a file download from a remote host via SFTP."""
    try:
        known_hosts_path = await _ensure_host_key(
            req.host, req.port, approve_host=_make_sftp_approve(req.approved_fingerprint)
        )
        connect_kwargs = _build_ssh_connect_kwargs(
            req.host, req.port, req.username, req.password, req.private_key, known_hosts_path
        )
        chunk_size = 65536  # 64 KB

        async def _stream_file():
            async with asyncssh.connect(**connect_kwargs) as conn:
                async with conn.start_sftp_client() as sftp:
                    async with sftp.open(req.path, 'rb') as remote_file:
                        while True:
                            chunk = await remote_file.read(chunk_size)
                            if not chunk:
                                break
                            yield chunk

        filename = req.path.rstrip('/').split('/')[-1]
        from starlette.responses import StreamingResponse  # pylint: disable=import-outside-toplevel
        return StreamingResponse(
            _stream_file(),
            media_type=_MIME_OCTET_STREAM,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("SFTP download error for %s: %s", req.host[:50], e)
        raise HTTPException(status_code=500, detail=_ERR_SFTP_FAILED) from e


@app.post(
    "/api/sftp/upload",
    summary="Upload a file via SFTP",
    responses={
        409: {"description": "Host key approval required"},
        500: {"description": _ERR_SFTP_FAILED},
    },
)
async def sftp_upload(  # pylint: disable=too-many-arguments,too-many-positional-arguments
    host: Annotated[str, Form(...)],
    username: Annotated[str, Form(...)],
    remote_path: Annotated[str, Form(...)],
    file: Annotated[UploadFile, File(...)],
    port: Annotated[int, Form()] = 22,
    password: Annotated[str | None, Form()] = None,
    private_key: Annotated[str | None, Form()] = None,
    approved_fingerprint: Annotated[str | None, Form()] = None,
):
    """Upload a file to a remote host via SFTP."""
    try:
        known_hosts_path = await _ensure_host_key(
            host, port, approve_host=_make_sftp_approve(approved_fingerprint)
        )
        connect_kwargs = _build_ssh_connect_kwargs(
            host, port, username, password, private_key, known_hosts_path
        )
        file_content = await file.read()
        remote_file_path = remote_path.rstrip('/') + '/' + file.filename

        async with asyncssh.connect(**connect_kwargs) as conn:
            async with conn.start_sftp_client() as sftp:
                async with sftp.open(remote_file_path, 'wb') as remote_file:
                    await remote_file.write(file_content)

        return {"success": True, "path": remote_file_path}
    except HTTPException:
        raise
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("SFTP upload error for %s: %s", host[:50], e)
        raise HTTPException(status_code=500, detail=_ERR_SFTP_FAILED) from e


@app.get("/api/wsl/discover", summary="Discover local WSL instances")
async def wsl_discover():
    """Discover locally-installed WSL instances by running wsl.exe."""
    try:
        process = await asyncio.create_subprocess_exec(
            _WSL_EXE, "-l", "-q",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )
        out, _ = await process.communicate()
        text = out.decode("utf-16le") if b"\x00" in out else out.decode("utf-8", errors="replace")
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        return {"wsl_instances": lines}
    except Exception:  # pylint: disable=broad-exception-caught
        return {"wsl_instances": []}


# Module-level set to hold references to fire-and-forget tasks so they
# are not garbage-collected before completion.
_pending_tasks: set = set()


def _tracked_task(coro):
    """Schedule a coroutine as an asyncio task, retaining a reference until done."""
    task = asyncio.create_task(coro)
    _pending_tasks.add(task)
    task.add_done_callback(_pending_tasks.discard)
    return task


def _parse_cpu_section(cpu_line: str, prev_idle: int, prev_total: int) -> tuple:
    """Parse the first line of /proc/stat and return (cpu_usage_float, new_idle, new_total)."""
    cpu_usage = 0
    idle, total = prev_idle, prev_total
    if cpu_line.startswith('cpu '):
        vals = [int(v) for v in cpu_line.split()[1:]]
        idle = vals[3] + vals[4]
        total = sum(vals)
        if prev_total > 0:
            diff_idle = idle - prev_idle
            diff_total = total - prev_total
            cpu_usage = (
                (1000 * (diff_total - diff_idle) / diff_total + 5) / 10
                if diff_total > 0 else 0
            )
    return cpu_usage, idle, total


def _parse_mem_section(mem_text: str) -> tuple:
    """Parse /proc/meminfo excerpt and return (mem_total_kb, mem_avail_kb, swap_total_kb, swap_free_kb)."""
    mem_total, mem_avail, swap_total, swap_free = 1, 0, 0, 0
    for line in mem_text.strip().split('\n'):
        if 'MemTotal' in line:
            mem_total = int(re.sub(_RE_NON_DIGIT, '', line))
        elif 'MemAvailable' in line:
            mem_avail = int(re.sub(_RE_NON_DIGIT, '', line))
        elif 'SwapTotal' in line:
            swap_total = int(re.sub(_RE_NON_DIGIT, '', line))
        elif 'SwapFree' in line:
            swap_free = int(re.sub(_RE_NON_DIGIT, '', line))
    return mem_total, mem_avail, swap_total, swap_free


def _parse_disk_section(disk_text: str) -> list:
    """Parse ``df -m`` output (NR>1 columns) and return a list of disk dicts."""
    disks = []
    _skip_fs = ('tmpfs', 'devtmpfs', 'overlay', 'shm')
    for d_line in disk_text.strip().split('\n'):
        tokens = d_line.strip().split()
        if len(tokens) < 6:
            continue
        fs = tokens[0]
        if fs in _skip_fs or fs.startswith('/dev/loop') or fs.startswith('squashfs'):
            continue
        try:
            disk_total = int(tokens[1])
            disk_used = int(tokens[2])
            disk_usage = (disk_used / disk_total) * 100 if disk_total > 0 else 0
            disks.append({"mount": tokens[5], "total_mb": disk_total, "used_mb": disk_used, "pct": disk_usage})
        except ValueError:
            pass
    return disks


def _parse_ssh_metrics(parts: list, prev_idle: int, prev_total: int) -> tuple:
    """Parse the four-section script output into a metrics payload.

    Returns (payload_dict, new_prev_idle, new_prev_total).
    """
    cpu_usage, prev_idle, prev_total = _parse_cpu_section(parts[0].strip(), prev_idle, prev_total)
    mem_total, mem_avail, swap_total, swap_free = _parse_mem_section(parts[1])
    disks = _parse_disk_section(parts[2])
    uptime_sec = float(parts[3].strip() or "0")
    ram_usage = (mem_total - mem_avail) / mem_total * 100 if mem_total > 0 else 0
    swap_usage = (swap_total - swap_free) / swap_total * 100 if swap_total > 0 else 0
    payload = {
        "type": "metrics",
        "cpu": min(max(cpu_usage, 0), 100),
        "ram_pct": ram_usage,
        "ram_total_mb": mem_total / 1024,
        "ram_used_mb": (mem_total - mem_avail) / 1024,
        "swap_pct": swap_usage,
        "swap_total_mb": swap_total / 1024,
        "swap_used_mb": (swap_total - swap_free) / 1024,
        "disks": disks,
        "uptime": uptime_sec,
    }
    return payload, prev_idle, prev_total


async def _dashboard_ws_approve_host(
    websocket: WebSocket,
    approved_fingerprint: str | None,
    h: str,
    p: int,
    fingerprint: str,
    _key: bytes,
) -> bool:
    """Approve a dashboard host key: auto-approve known fingerprint or prompt the browser."""
    if approved_fingerprint and approved_fingerprint == fingerprint:
        return True
    await websocket.send_json({
        "type": "host_key_approval",
        "host": h,
        "port": p,
        "fingerprint": fingerprint,
    })
    async with asyncio.timeout(60):
         return await _ws_wait_for_host_key_response(websocket)


async def _ssh_dashboard_connect(websocket: WebSocket, config: dict) -> None:
    """Parse config, verify host key, connect and start the metrics loop."""
    ssh_host = config.get("host")
    port = int(config.get("port", 22))
    username = config.get("username")
    password = config.get("password")
    private_key = config.get("private_key")
    approved_fingerprint = config.get("approved_fingerprint")

    async def _ws_approve_host(h: str, p: int, fingerprint: str, _key: bytes) -> bool:
        return await _dashboard_ws_approve_host(websocket, approved_fingerprint, h, p, fingerprint, _key)

    try:
        known_hosts_path = await _ensure_host_key(ssh_host, port, approve_host=_ws_approve_host)
    except RuntimeError as exc:
        await websocket.send_json({"error": str(exc)})
        await websocket.close()
        return
    except Exception as e:  # pylint: disable=broad-exception-caught
        await websocket.send_json({"error": f"Host verification failed: {e}"})
        await websocket.close()
        return

    connect_kwargs = _build_ssh_connect_kwargs(ssh_host, port, username, password, private_key, known_hosts_path)
    try:
        async with asyncssh.connect(**connect_kwargs) as conn:
            await _run_metrics_loop(websocket, conn)
    except WebSocketDisconnect:
        raise
    except Exception as e:  # pylint: disable=broad-exception-caught
        await websocket.send_json({"error": f"SSH connection error: {e}"})
        await websocket.close()


@app.websocket("/api/ssh/dashboard")
async def ssh_dashboard(websocket: WebSocket):
    """WebSocket endpoint: real-time SSH server metrics dashboard."""
    if not await _ws_check_origin(websocket):
        return

    await websocket.accept()
    try:
        data = await websocket.receive_text()
        config = json.loads(data)
        await _ssh_dashboard_connect(websocket, config)
    except WebSocketDisconnect:
        pass
    except Exception as e:  # pylint: disable=broad-exception-caught
        try:
            await websocket.send_json({"error": f"Connection lost: {e}"})
            await websocket.close()
        except Exception:  # pylint: disable=broad-exception-caught
            logger.debug("ssh_dashboard: failed to close websocket after error", exc_info=True)


def _make_pty_read_handler(fd: int, loop, websocket: WebSocket):
    """Return a callback that reads from *fd* and forwards data over *websocket*."""
    def on_pty_read():
        try:
            data = os.read(fd, 8192)
            if data:
                _tracked_task(websocket.send_text(data.decode("utf-8", errors="replace")))
            else:
                loop.remove_reader(fd)
                _tracked_task(websocket.close())
        except OSError:
            loop.remove_reader(fd)
            _tracked_task(websocket.close())
    return on_pty_read


def _apply_terminal_resize(fd: int, data: str, resize_pattern) -> bool:
    """Apply a terminal resize if *data* matches the resize escape sequence.

    Returns True if the message was a resize and was handled, False otherwise.
    """
    match = resize_pattern.match(data)
    if not match:
        return False
    cols = int(match.group(1))
    rows = int(match.group(2))
    winsize = struct.pack("HHHH", rows, cols, 0, 0)
    fcntl.ioctl(fd, termios.TIOCSWINSZ, winsize)
    return True


def _parse_distro_from_config(config_raw: str) -> str | None:
    """Parse and validate the distro name from a JSON config string."""
    try:
        config = json.loads(config_raw)
        distro = config.get("distro")
        if distro and not _DISTRO_NAME_RE.match(distro):
            logger.warning("local_terminal: rejected invalid distro name %r", distro)
            return None
        return distro
    except Exception:  # pylint: disable=broad-exception-caught
        logger.debug("local_terminal: failed to parse config JSON, proceeding with distro=None")
        return None


def _exec_pty_child(distro: str | None) -> None:  # pragma: no cover — runs in forked child
    """Replace the child process image with the appropriate shell or WSL distro."""
    current_distro = os.environ.get("WSL_DISTRO_NAME", "")
    if distro and distro != current_distro:
        # Note: wsl.exe across PTY interop might hang in certain builds,
        # but we allow it for cross-distro attempts.
        os.execvp(_WSL_EXE, [_WSL_EXE, "-d", distro])  # nosec B606
    else:
        shell = os.environ.get("SHELL", "/bin/bash")
        os.execvp(shell, [shell])  # nosec B606


async def _run_local_pty_loop(websocket: WebSocket, fd: int) -> None:
    """Run the PTY read/write loop until the WebSocket disconnects."""
    loop = asyncio.get_running_loop()
    loop.add_reader(fd, _make_pty_read_handler(fd, loop, websocket))
    resize_pattern = re.compile(r"^\x1b\[resize;(\d+);(\d+)m$")
    try:
        while True:
            data = await websocket.receive_text()
            if not _apply_terminal_resize(fd, data, resize_pattern):
                os.write(fd, data.encode("utf-8"))
    except WebSocketDisconnect:
        pass
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("local_terminal: unexpected error: %s", e)
    finally:
        try:
            loop.remove_reader(fd)
            os.close(fd)
        except Exception:  # pylint: disable=broad-exception-caught
            logger.debug("local_terminal: error during cleanup", exc_info=True)


@app.websocket("/api/local/terminal")
async def local_terminal(websocket: WebSocket):
    """WebSocket endpoint: local PTY terminal (Linux/macOS only)."""
    if not await _ws_check_origin(websocket):
        return

    if not _pty_available:
        await websocket.close(code=1008, reason="Local terminal is not supported on this platform")
        return

    await websocket.accept()

    config_raw = await websocket.receive_text()
    distro = _parse_distro_from_config(config_raw)

    pid, fd = pty.fork()
    if pid == 0:
        _exec_pty_child(distro)

    await _run_local_pty_loop(websocket, fd)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
