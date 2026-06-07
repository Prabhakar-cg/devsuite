"""
routes/convert.py — Server-side file format conversion (/api/convert).
"""
import json
import urllib.parse
from typing import Annotated

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import Response

from deps import (
    MAX_UPLOAD_SIZE,
    _MIME_OCTET_STREAM,
    _OPENPYXL_MISSING,
    _XLSX_MEDIA_TYPE,
)

router = APIRouter()


# ─── Conversion helpers ───────────────────────────────────────────────────────

def _conv_xlsx_to_csv_json(content: bytes, target_format: str) -> Response:
    """Convert XLSX bytes to CSV or JSON and return the appropriate Response."""
    import io  # noqa: PLC0415
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as e:
        raise HTTPException(status_code=503, detail=_OPENPYXL_MISSING) from e
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Spreadsheet is empty")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [
        dict(zip(headers, [str(v) if v is not None else "" for v in row]))
        for row in rows[1:]
    ]
    if target_format == "csv":
        import csv as csv_mod  # noqa: PLC0415
        buf = io.StringIO()
        writer = csv_mod.DictWriter(buf, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": 'attachment; filename="converted.csv"'})
    return Response(content=json.dumps(data, indent=2), media_type="application/json",
                    headers={"Content-Disposition": 'attachment; filename="converted.json"'})


def _conv_csv_to_xlsx(content: bytes) -> Response:
    """Convert CSV bytes to XLSX and return the appropriate Response."""
    import io  # noqa: PLC0415
    import csv as csv_mod  # noqa: PLC0415
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as e:
        raise HTTPException(status_code=503, detail=_OPENPYXL_MISSING) from e
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv_mod.reader(io.StringIO(text))
    rows = list(reader)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="converted.xlsx"'},
    )


def _conv_json_to_xlsx(content: bytes) -> Response:
    """Convert JSON bytes to XLSX and return the appropriate Response."""
    import io  # noqa: PLC0415
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as e:
        raise HTTPException(status_code=503, detail=_OPENPYXL_MISSING) from e
    data = json.loads(content)
    if not isinstance(data, list):
        data = [data]
    wb = openpyxl.Workbook()
    ws = wb.active
    if data:
        if not isinstance(data[0], dict):
            raise HTTPException(status_code=400, detail="JSON must be an array of objects for XLSX conversion")
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([str(row.get(h, "")) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="converted.xlsx"'},
    )


def _conv_pdf_to_txt(content: bytes) -> Response:
    """Extract text from PDF bytes and return a plain-text Response."""
    import io  # noqa: PLC0415
    try:
        import pypdf  # noqa: PLC0415
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail="pypdf is not installed. Run: pip install pypdf"
        ) from e
    reader = pypdf.PdfReader(io.BytesIO(content))
    pages_text = [page.extract_text() or "" for page in reader.pages]
    full_text = "\n\n".join(pages_text)
    return Response(content=full_text, media_type="text/plain",
                    headers={"Content-Disposition": 'attachment; filename="converted.txt"'})


def _conv_docx_to_txt(content: bytes) -> Response:
    """Extract text from DOCX bytes and return a plain-text Response."""
    import io  # noqa: PLC0415
    try:
        import docx  # noqa: PLC0415
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail="python-docx is not installed. Run: pip install python-docx"
        ) from e
    doc = docx.Document(io.BytesIO(content))
    text = "\n".join(para.text for para in doc.paragraphs)
    return Response(content=text, media_type="text/plain",
                    headers={"Content-Disposition": 'attachment; filename="converted.txt"'})


def _source_to_html(src_ext: str, content: bytes) -> str:
    """Convert document bytes to a raw HTML string for PDF rendering."""
    import io  # noqa: PLC0415
    if src_ext in ("docx", "doc"):
        try:
            import mammoth  # noqa: PLC0415
        except ImportError as e:
            raise HTTPException(
                status_code=503,
                detail="mammoth is not installed. Run: pip install mammoth"
            ) from e
        return mammoth.convert_to_html(io.BytesIO(content)).value
    if src_ext in ("md", "markdown"):
        import html as _html_mod  # noqa: PLC0415
        try:
            import markdown as md_lib  # noqa: PLC0415
            return md_lib.markdown(
                content.decode("utf-8", errors="replace"),
                extensions=["tables", "fenced_code"],
            )
        except ImportError:
            return "<pre>" + _html_mod.escape(content.decode("utf-8", errors="replace")) + "</pre>"
    return content.decode("utf-8", errors="replace")  # html / htm


def _conv_any_to_pdf(src_ext: str, content: bytes) -> Response:
    """Convert DOCX/DOC/HTML/MD/Markdown bytes to PDF using weasyprint."""
    try:
        import weasyprint  # noqa: PLC0415
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail="weasyprint is not installed. Run: pip install weasyprint"
        ) from e

    raw_html = _source_to_html(src_ext, content)

    full_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Georgia, 'Times New Roman', serif; font-size: 12pt; line-height: 1.6;
          margin: 2cm; color: #111; }}
  h1,h2,h3,h4,h5,h6 {{ font-family: Arial, Helvetica, sans-serif; margin-top: 1.2em; }}
  h1 {{ font-size: 22pt; }} h2 {{ font-size: 17pt; }} h3 {{ font-size: 14pt; }}
  p {{ margin: 0.5em 0 0.8em; }}
  table {{ border-collapse: collapse; width: 100%; margin: 1em 0; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; text-align: left; }}
  th {{ background: #f0f0f0; font-weight: bold; }}
  pre, code {{ font-family: 'Courier New', monospace; font-size: 10pt;
               background: #f5f5f5; padding: 0.2em 0.4em; border-radius: 3px; }}
  pre {{ padding: 0.8em; white-space: pre-wrap; word-break: break-word; }}
  img {{ max-width: 100%; height: auto; }}
  a {{ color: #1a56db; }}
  @page {{ margin: 2cm; }}
</style>
</head>
<body>{raw_html}</body>
</html>"""

    try:
        from weasyprint import default_url_fetcher as _default_url_fetcher  # noqa: PLC0415
    except ImportError:
        _default_url_fetcher = None

    def _safe_url_fetcher(url: str):
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in ('', 'data'):
            raise ValueError(
                f"Blocked disallowed URL scheme '{parsed.scheme}' in PDF conversion: {url!r}"
            )
        if _default_url_fetcher is not None:
            return _default_url_fetcher(url)
        raise ValueError(f"No URL fetcher available for: {url!r}")

    pdf_bytes = weasyprint.HTML(string=full_html, url_fetcher=_safe_url_fetcher).write_pdf()
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": 'attachment; filename="converted.pdf"'})


def _check_content_length_header(cl: str | None) -> None:
    """Raise HTTPException 400/413 if the Content-Length header is invalid or too large."""
    if not cl:
        return
    try:
        cl_int = int(cl)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid Content-Length header") from exc
    if cl_int < 0:
        raise HTTPException(status_code=400, detail="Invalid Content-Length header")
    if cl_int > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Upload too large (limit {MAX_UPLOAD_SIZE // 1024 // 1024} MB)",
        )


# ─── Route ────────────────────────────────────────────────────────────────────

@router.post(
    "/api/convert",
    summary="Convert a file from one format to another (server-side)",
    responses={
        400: {"description": "Invalid content-length or unsupported conversion"},
        413: {"description": "Upload too large"},
        503: {"description": "Required conversion library not installed"},
    },
)
async def convert_file(
    request: Request,
    file: Annotated[UploadFile, File(...)],
    target_format: Annotated[str, Form(...)],
):
    """Server-side file format conversion endpoint."""
    target_format = target_format.lower().strip()
    original_name = (file.filename or "file").lower()
    src_ext = original_name.rsplit(".", 1)[-1] if "." in original_name else ""

    _check_content_length_header(request.headers.get("content-length"))
    content = await file.read(MAX_UPLOAD_SIZE + 1)
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Upload too large (limit {MAX_UPLOAD_SIZE // 1024 // 1024} MB)",
        )

    if src_ext == "xlsx" and target_format in ("csv", "json"):
        return _conv_xlsx_to_csv_json(content, target_format)
    if src_ext == "csv" and target_format == "xlsx":
        return _conv_csv_to_xlsx(content)
    if src_ext == "json" and target_format == "xlsx":
        return _conv_json_to_xlsx(content)
    if src_ext == "pdf" and target_format == "txt":
        return _conv_pdf_to_txt(content)
    if src_ext == "docx" and target_format == "txt":
        return _conv_docx_to_txt(content)
    if target_format == "pdf" and src_ext in ("docx", "doc", "html", "htm", "md", "markdown"):
        return _conv_any_to_pdf(src_ext, content)
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported server-side conversion: {src_ext} → {target_format}"
    )
